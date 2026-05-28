"""Управление AI: системный промпт, база знаний, быстрые вопросы + импорт XLSX."""

import json
import os
import hashlib
import requests
import openpyxl
import psycopg2
import boto3
import uuid
import base64
from io import BytesIO

XLSX_URL = 'https://cdn.poehali.dev/projects/73fc8821-802d-4489-8ce7-ef196540fbf0/bucket/c04d82f2-0303-48f7-b069-9c96c191ffb8.xlsx'
SCHEMA = os.environ.get('MAIN_DB_SCHEMA', 'public')
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', '').strip() or 'Sdauxbasstre228'

CORS = {
    'Access-Control-Allow-Origin': '*',
    'Access-Control-Allow-Methods': 'GET, POST, PUT, DELETE, OPTIONS',
    'Access-Control-Allow-Headers': 'Content-Type, X-Admin-Token, Authorization',
}

def resp(status, body):
    return {'statusCode': status, 'headers': {**CORS, 'Content-Type': 'application/json'}, 'body': json.dumps(body, ensure_ascii=False)}

def get_conn():
    return psycopg2.connect(os.environ['DATABASE_URL'])

def _extract_token(headers: dict) -> str:
    raw = (headers.get('x-authorization') or headers.get('X-Authorization')
           or headers.get('x-admin-token') or headers.get('X-Admin-Token') or '')
    return raw.removeprefix('Bearer ').strip()

def get_wl_id(headers: dict):
    """Возвращает wl_manager_id если запрос от WL-клиента, иначе None.
    WL-клиент = сессия с session_type='wl_manager' в user_sessions."""
    token = _extract_token(headers)
    if not token:
        return None
    # Мастер-пароль → не WL
    if ADMIN_PASSWORD and hashlib.sha256(token.encode()).hexdigest() == hashlib.sha256(ADMIN_PASSWORD.encode()).hexdigest():
        return None
    try:
        conn = get_conn(); cur = conn.cursor()
        cur.execute(
            f"SELECT user_id FROM {SCHEMA}.user_sessions WHERE token=%s AND expires_at>NOW() AND session_type='wl_manager'",
            (token,)
        )
        row = cur.fetchone(); cur.close(); conn.close()
        return row[0] if row else None
    except Exception:
        return None

def check_auth(headers: dict) -> bool:
    token = _extract_token(headers)
    if not token:
        return False
    # 1. Мастер-пароль
    if ADMIN_PASSWORD and hashlib.sha256(token.encode()).hexdigest() == hashlib.sha256(ADMIN_PASSWORD.encode()).hexdigest():
        return True
    # 2. Сессионный токен (любой пользователь или WL-менеджер)
    try:
        conn = get_conn(); cur = conn.cursor()
        cur.execute(
            f"SELECT s.id FROM {SCHEMA}.user_sessions s WHERE s.token=%s AND s.expires_at>NOW()",
            (token,)
        )
        row = cur.fetchone(); cur.close(); conn.close()
        return row is not None
    except Exception:
        return False

def get_company_id_from_token(headers: dict):
    """Возвращает (user_id, role) по X-Authorization токену сессии, или (None, None)."""
    token = _extract_token(headers)
    if not token:
        return None, None
    conn = get_conn(); cur = conn.cursor()
    cur.execute(
        f"SELECT u.id, u.role FROM {SCHEMA}.user_sessions s JOIN {SCHEMA}.users u ON u.id=s.user_id WHERE s.token=%s AND s.expires_at>NOW()",
        (token,)
    )
    row = cur.fetchone(); cur.close(); conn.close()
    if not row:
        return None, None
    return row[0], row[1]

def load_xlsx():
    r = requests.get(XLSX_URL, timeout=15)
    wb = openpyxl.load_workbook(BytesIO(r.content), data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    hdrs = [str(h).strip().lower() if h else '' for h in rows[0]]
    items = []
    for row in rows[1:]:
        if not any(row):
            continue
        item = {}
        for i, h in enumerate(hdrs):
            val = row[i] if i < len(row) else None
            item[h] = str(val).strip() if val is not None else ''
        items.append(item)
    return items, hdrs


_COMPLEX_STOP_WORDS = [
    'лента', 'двухуровн', 'керамогран', 'вентил', 'блок питани', 'парящ',
    'теневой', 'тенев', 'краб', 'плитк', 'диффузор', 'дифузор', 'без монт', 'вклейк',
]

def _save_complex_exceptions(conn, cur, synonyms_str: str, name: str):
    """Если синоним/имя покрывает стоп-слово — добавляем исключение в БД."""
    all_texts = [name] + [s.strip() for s in (synonyms_str or '').split(',') if s.strip()]
    for text in all_texts:
        tl = text.lower()
        for cw in _COMPLEX_STOP_WORDS:
            if cw in tl:
                cur.execute(
                    f"INSERT INTO {SCHEMA}.complex_word_exceptions (word) VALUES (%s) ON CONFLICT DO NOTHING",
                    (cw,)
                )
    conn.commit()

def _ensure_wl_initialized(conn, cur, wl_id: int, resource: str):
    """Если WL-клиент первый раз открывает FAQ или questions — копируем глобальные данные."""
    flag_col = 'faq_initialized' if resource == 'faq' else 'questions_initialized'
    cur.execute(
        f"INSERT INTO {SCHEMA}.wl_data_initialized (wl_manager_id) VALUES (%s) ON CONFLICT DO NOTHING",
        (wl_id,)
    )
    cur.execute(
        f"SELECT {flag_col} FROM {SCHEMA}.wl_data_initialized WHERE wl_manager_id=%s",
        (wl_id,)
    )
    row = cur.fetchone()
    if row and row[0]:
        return  # уже инициализировано
    # Копируем глобальные данные
    if resource == 'faq':
        cur.execute(f"SELECT title, content, used FROM {SCHEMA}.faq_items ORDER BY id")
        global_rows = cur.fetchall()
        for gr in global_rows:
            cur.execute(
                f"INSERT INTO {SCHEMA}.wl_faq_items (wl_manager_id, title, content, used) VALUES (%s,%s,%s,%s)",
                (wl_id, gr[0], gr[1], gr[2])
            )
        cur.execute(
            f"UPDATE {SCHEMA}.wl_data_initialized SET faq_initialized=TRUE WHERE wl_manager_id=%s",
            (wl_id,)
        )
    else:
        cur.execute(f"SELECT pattern, answer, active FROM {SCHEMA}.ai_quick_questions ORDER BY id")
        global_rows = cur.fetchall()
        for gr in global_rows:
            cur.execute(
                f"INSERT INTO {SCHEMA}.wl_quick_questions (wl_manager_id, pattern, answer, active) VALUES (%s,%s,%s,%s)",
                (wl_id, gr[0], gr[1], gr[2])
            )
        cur.execute(
            f"UPDATE {SCHEMA}.wl_data_initialized SET questions_initialized=TRUE WHERE wl_manager_id=%s",
            (wl_id,)
        )
    conn.commit()


def handler(event: dict, context) -> dict:
    """Управление AI-конфигурацией: промпт, база знаний, быстрые вопросы, импорт XLSX.
    Маршрутизация через query-параметр: ?r=prompt|faq|questions|prices|login
    WL-клиенты получают свои данные через override-таблицы, мастер работает с глобальными.
    """
    if event.get('httpMethod') == 'OPTIONS':
        return {'statusCode': 200, 'headers': CORS, 'body': ''}

    method = event.get('httpMethod', 'GET')
    hdrs = event.get('headers') or {}
    qs = event.get('queryStringParameters') or {}
    r = qs.get('r', '')
    body_str = event.get('body') or '{}'

    # Определяем: мастер или WL-клиент
    wl_id = get_wl_id(hdrs)
    is_master = (wl_id is None) and check_auth(hdrs)

    # --- POST ?r=login
    if r == 'login' and method == 'POST':
        body = json.loads(body_str)
        password = body.get('password', '')
        pw_hash = hashlib.sha256(password.encode()).hexdigest()
        if pw_hash == hashlib.sha256(ADMIN_PASSWORD.encode()).hexdigest():
            return resp(200, {'token': password})
        return resp(401, {'error': 'Неверный пароль'})

    # ══════════════════════════════════════════════════════════════
    # ПРОМПТ
    # ══════════════════════════════════════════════════════════════

    # --- GET ?r=prompt
    if r == 'prompt' and method == 'GET':
        conn = get_conn(); cur = conn.cursor()
        if wl_id:
            # WL: сначала своё переопределение
            cur.execute(f"SELECT content, updated_at FROM {SCHEMA}.wl_prompt_override WHERE wl_manager_id=%s", (wl_id,))
            row = cur.fetchone()
            if row:
                cur.close(); conn.close()
                return resp(200, {'id': wl_id, 'content': row[0], 'updated_at': str(row[1])})
        # Глобальный промпт (мастер или WL без своего)
        cur.execute(f"SELECT id, content, updated_at FROM {SCHEMA}.ai_system_prompt ORDER BY id LIMIT 1")
        row = cur.fetchone()
        cur.close(); conn.close()
        if not row:
            return resp(404, {'error': 'Not found'})
        return resp(200, {'id': row[0], 'content': row[1], 'updated_at': str(row[2])})

    # --- PUT ?r=prompt
    if r == 'prompt' and method == 'PUT':
        if not check_auth(hdrs):
            return resp(401, {'error': 'Unauthorized'})
        body = json.loads(body_str)
        content = body.get('content', '').strip()
        if not content:
            return resp(400, {'error': 'content required'})
        conn = get_conn(); cur = conn.cursor()
        if wl_id:
            # WL: сохраняем в своё переопределение
            cur.execute(
                f"INSERT INTO {SCHEMA}.wl_prompt_override (wl_manager_id, content, updated_at) VALUES (%s,%s,now()) ON CONFLICT (wl_manager_id) DO UPDATE SET content=%s, updated_at=now()",
                (wl_id, content, content)
            )
        else:
            # Мастер: обновляет глобальный
            cur.execute(
                f"UPDATE {SCHEMA}.ai_system_prompt SET content=%s, updated_at=now() WHERE id=(SELECT id FROM {SCHEMA}.ai_system_prompt ORDER BY id LIMIT 1)",
                (content,)
            )
        conn.commit(); cur.close(); conn.close()
        return resp(200, {'ok': True})

    # ══════════════════════════════════════════════════════════════
    # FAQ
    # ══════════════════════════════════════════════════════════════

    # --- GET ?r=faq
    if r == 'faq' and method == 'GET':
        conn = get_conn(); cur = conn.cursor()
        if wl_id:
            _ensure_wl_initialized(conn, cur, wl_id, 'faq')
            cur.execute(f"SELECT id, title, content, used, created_at, COALESCE(images, '[]'::jsonb), COALESCE(items, '[]'::jsonb) FROM {SCHEMA}.wl_faq_items WHERE wl_manager_id=%s ORDER BY id", (wl_id,))
        else:
            cur.execute(f"SELECT id, title, content, used, created_at, COALESCE(images, '[]'::jsonb), COALESCE(items, '[]'::jsonb) FROM {SCHEMA}.faq_items ORDER BY id")
        rows = cur.fetchall()
        cur.close(); conn.close()
        def parse_json(v): return v if isinstance(v, list) else (json.loads(v) if v else [])
        return resp(200, {'items': [{'id': row[0], 'title': row[1], 'content': row[2], 'used': row[3], 'created_at': str(row[4]), 'images': parse_json(row[5]), 'items': parse_json(row[6])} for row in rows]})

    # --- POST ?r=faq
    if r == 'faq' and method == 'POST':
        if not check_auth(hdrs):
            return resp(401, {'error': 'Unauthorized'})
        body = json.loads(body_str)
        title = body.get('title', '').strip()
        content = body.get('content', '').strip()
        images = json.dumps(body.get('images', []))
        items  = json.dumps(body.get('items', []))
        if not title:
            return resp(400, {'error': 'title required'})
        conn = get_conn(); cur = conn.cursor()
        if wl_id:
            _ensure_wl_initialized(conn, cur, wl_id, 'faq')
            cur.execute(f"INSERT INTO {SCHEMA}.wl_faq_items (wl_manager_id, title, content, used, images, items) VALUES (%s,%s,%s,true,%s,%s) RETURNING id", (wl_id, title, content, images, items))
        else:
            cur.execute(f"INSERT INTO {SCHEMA}.faq_items (title, content, used, images, items) VALUES (%s, %s, true, %s, %s) RETURNING id", (title, content, images, items))
        new_id = cur.fetchone()[0]
        conn.commit(); cur.close(); conn.close()
        return resp(200, {'id': new_id, 'ok': True})

    # --- PUT ?r=faq&id=X
    if r == 'faq' and method == 'PUT':
        if not check_auth(hdrs):
            return resp(401, {'error': 'Unauthorized'})
        faq_id = int(qs.get('id', '0'))
        body = json.loads(body_str)
        images = json.dumps(body.get('images', []))
        items  = json.dumps(body.get('items', []))
        conn = get_conn(); cur = conn.cursor()
        if wl_id:
            cur.execute(
                f"UPDATE {SCHEMA}.wl_faq_items SET title=%s, content=%s, used=%s, images=%s, items=%s WHERE id=%s AND wl_manager_id=%s",
                (body.get('title',''), body.get('content',''), body.get('used', True), images, items, faq_id, wl_id)
            )
        else:
            cur.execute(
                f"UPDATE {SCHEMA}.faq_items SET title=%s, content=%s, used=%s, images=%s, items=%s WHERE id=%s",
                (body.get('title',''), body.get('content',''), body.get('used', True), images, items, faq_id)
            )
        conn.commit(); cur.close(); conn.close()
        return resp(200, {'ok': True})

    # --- DELETE ?r=faq&id=X
    if r == 'faq' and method == 'DELETE':
        if not check_auth(hdrs):
            return resp(401, {'error': 'Unauthorized'})
        faq_id = int(qs.get('id', '0'))
        conn = get_conn(); cur = conn.cursor()
        if wl_id:
            cur.execute(f"DELETE FROM {SCHEMA}.wl_faq_items WHERE id=%s AND wl_manager_id=%s", (faq_id, wl_id))
        else:
            cur.execute(f"DELETE FROM {SCHEMA}.faq_items WHERE id = %s", (faq_id,))
        conn.commit(); cur.close(); conn.close()
        return resp(200, {'ok': True})

    # --- POST ?r=faq-migrate-items  — разовая миграция: content → items
    if r == 'faq-migrate-items' and method == 'POST':
        if not check_auth(hdrs):
            return resp(401, {'error': 'Unauthorized'})
        import re as _re

        def parse_content_to_items(content: str) -> list:
            """Разбивает большой текст на товары по двойному переносу строки.
            Для каждого блока ищет название из строк Товар:/Наименование:."""
            if not content or not content.strip():
                return []
            blocks = _re.split(r'\n\s*\n', content.strip())
            items = []
            for block in blocks:
                block = block.strip()
                if not block:
                    continue
                # Ищем название: строки начинающиеся с Товар: или Наименование:
                name = None
                for line in block.splitlines():
                    m = _re.match(r'^(?:Товар|Наименование)\s*:\s*(.+)', line.strip(), _re.IGNORECASE)
                    if m:
                        name = m.group(1).strip().rstrip('.')
                        break
                # Если не нашли — берём первую строку как название
                if not name:
                    first_line = block.splitlines()[0].strip()
                    name = first_line[:80]
                items.append({
                    'id': _re.sub(r'[^a-z0-9]', '', name.lower())[:12] + str(abs(hash(name)) % 10000),
                    'name': name,
                    'description': block,
                    'image_url': '',
                })
            return items

        conn = get_conn(); cur = conn.cursor()
        cur.execute(f"SELECT id, content, items FROM {SCHEMA}.faq_items WHERE items = '[]'::jsonb OR items IS NULL")
        rows = cur.fetchall()
        updated = 0
        for row in rows:
            faq_id, content, existing_items = row
            if existing_items and existing_items != [] and len(existing_items) > 0:
                continue
            parsed = parse_content_to_items(content or '')
            if parsed:
                cur.execute(f"UPDATE {SCHEMA}.faq_items SET items = %s WHERE id = %s",
                            (json.dumps(parsed, ensure_ascii=False), faq_id))
                updated += 1
        conn.commit(); cur.close(); conn.close()
        return resp(200, {'ok': True, 'updated': updated})

    # --- POST ?r=faq-search-image  — поиск картинок через Tavily (до 5 штук), сохранение в S3
    if r == 'faq-search-image' and method == 'POST':
        if not check_auth(hdrs):
            return resp(401, {'error': 'Unauthorized'})
        body_data = json.loads(body_str)
        query = (body_data.get('query') or '').strip()
        limit = min(int(body_data.get('limit', 5)), 5)
        # URL которые уже есть или были отклонены пользователем — не возвращать
        exclude_urls = set(body_data.get('exclude_urls') or [])
        if not query:
            return resp(400, {'error': 'query required'})

        tavily_key = os.environ.get('TAVILY_API_KEY', '')
        if not tavily_key:
            return resp(500, {'error': 'TAVILY_API_KEY не настроен'})

        SPAM = ('passport', 'document', 'avito', 'ozon', 'wildberries',
                'instagram', 'facebook', 'vk.com', 'youtube', 'tiktok')

        def tavily_search(search_query):
            tv = requests.post(
                'https://api.tavily.com/search',
                json={
                    'api_key': tavily_key,
                    'query': search_query,
                    'search_depth': 'basic',
                    'max_results': 8,
                    'include_images': True,
                    'include_image_descriptions': True,
                },
                timeout=20,
            )
            if tv.status_code != 200:
                return []
            raw = tv.json().get('images', [])
            result = []
            for img in raw:
                img_url = img.get('url') if isinstance(img, dict) else str(img)
                if not img_url or not img_url.startswith('http'):
                    continue
                if any(s in img_url.lower() for s in SPAM):
                    continue
                if img_url in exclude_urls:
                    continue
                result.append(img_url)
            return result

        try:
            # 1. Пробуем несколько разных формулировок запроса
            query_variants = [
                f"натяжной потолок {query} фото интерьер",
                f"{query} потолок дизайн фото",
                f"натяжной потолок {query} вид",
            ]
            candidate_urls = []
            for variant in query_variants:
                if len(candidate_urls) >= limit * 2:
                    break
                new_urls = tavily_search(variant)
                # Добавляем только те что ещё не встречались
                for u in new_urls:
                    if u not in candidate_urls and u not in exclude_urls:
                        candidate_urls.append(u)

            if not candidate_urls:
                return resp(404, {'error': 'Картинки не найдены'})

            # 2. Скачиваем и сохраняем в S3, запоминаем source→cdn маппинг
            s3 = boto3.client('s3', endpoint_url='https://bucket.poehali.dev',
                              aws_access_key_id=os.environ['AWS_ACCESS_KEY_ID'],
                              aws_secret_access_key=os.environ['AWS_SECRET_ACCESS_KEY'])
            results = []  # [{cdn, source}]
            ua = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            for img_url in candidate_urls:
                if len(results) >= limit:
                    break
                try:
                    dl = requests.get(img_url, headers={'User-Agent': ua}, timeout=8, stream=False)
                    if dl.status_code != 200 or len(dl.content) < 2000:
                        continue
                    ct = dl.headers.get('Content-Type', 'image/jpeg').split(';')[0].strip()
                    if not ct.startswith('image/'):
                        continue
                    ext = ct.split('/')[-1].replace('jpeg', 'jpg').replace('png', 'png') or 'jpg'
                    file_key = f'faq-images/{uuid.uuid4().hex}.{ext}'
                    s3.put_object(Bucket='files', Key=file_key, Body=dl.content, ContentType=ct)
                    cdn_url = f"https://cdn.poehali.dev/projects/{os.environ['AWS_ACCESS_KEY_ID']}/bucket/{file_key}"
                    results.append({'cdn': cdn_url, 'source': img_url})
                except Exception:
                    continue

            if not results:
                return resp(404, {'error': 'Не удалось сохранить ни одну картинку'})

            return resp(200, {
                'urls': [r['cdn'] for r in results],
                'sources': [r['source'] for r in results],
                'count': len(results),
            })

        except Exception as e:
            return resp(500, {'error': str(e)})

    # --- POST ?r=faq-enrich-product  — сбор данных о товаре из интернета через Tavily + AI
    if r == 'faq-enrich-product' and method == 'POST':
        print('faq-enrich-product called, method=', method)
        auth_ok = check_auth(hdrs)
        print('auth_ok=', auth_ok)
        if not auth_ok:
            return resp(401, {'error': 'Unauthorized'})
        try:
            body = json.loads(body_str)
        except Exception as ex:
            print('body parse error:', ex, 'body_str=', body_str[:200])
            return resp(400, {'error': 'bad json'})
        product_name = (body.get('name') or '').strip()
        current_description = (body.get('description') or '').strip()
        category_name = (body.get('category') or '').strip()
        print('product_name=', product_name, 'desc_len=', len(current_description))
        if not product_name:
            return resp(400, {'error': 'Укажите название товара'})

        tavily_key = os.environ.get('TAVILY_API_KEY', '')
        openrouter_key = os.environ.get('OPENROUTER_API_KEY_2', '')
        print('tavily_key present=', bool(tavily_key), 'openrouter_key present=', bool(openrouter_key))
        if not tavily_key:
            return resp(500, {'error': 'TAVILY_API_KEY не настроен'})
        if not openrouter_key:
            return resp(500, {'error': 'OPENROUTER_API_KEY_2 не настроен'})

        try:
            print('entering try block')
            # 1. Ищем информацию о товаре через Tavily
            context_parts = []
            for query in [
                f'{product_name} натяжной потолок характеристики цена',
                f'{product_name} {category_name} описание производитель'.strip(),
            ]:
                tv = requests.post(
                    'https://api.tavily.com/search',
                    json={
                        'api_key': tavily_key,
                        'query': query,
                        'search_depth': 'basic',
                        'max_results': 4,
                        'include_answer': True,
                    },
                    timeout=15,
                )
                if tv.status_code == 200:
                    data = tv.json()
                    if data.get('answer'):
                        context_parts.append(data['answer'])
                    for res in (data.get('results') or [])[:3]:
                        snippet = res.get('content') or res.get('snippet') or ''
                        if snippet:
                            context_parts.append(snippet[:600])

            web_context = '\n\n'.join(context_parts[:6]) if context_parts else ''

            # 2. AI генерирует описание в строгом формате
            system_prompt = (
                'Ты помощник для компании по натяжным потолкам. '
                'На основе информации составь описание товара для базы знаний AI-агента. '
                'Отвечай СТРОГО в следующем формате — каждый пункт с новой строки, без лишних слов:\n'
                'Товар: <название>\n'
                'Категория: <категория>\n'
                'Цена: <цена за единицу, если известна>\n'
                'Материал: <материал>\n'
                'Цвет: <цвет или цвета>\n'
                'Размер: <размеры, длина, ширина, высота профиля и т.п.>\n'
                'Применение: <для чего используется, кратко>\n'
                'Где используется: <в каких помещениях или условиях>\n\n'
                'Если какой-то параметр неизвестен — пиши «уточняется». '
                'Не добавляй лишних абзацев, пояснений и заголовков. '
                'Отвечай только на русском языке.'
            )
            user_prompt = (
                f'Товар: {product_name}\n'
                + (f'Категория: {category_name}\n' if category_name else '')
                + (f'Текущее описание: {current_description}\n' if current_description else '')
                + (f'\nИнформация из интернета:\n{web_context}' if web_context else '\nИнформации из интернета не найдено — составь типичное описание для этого товара.')
            )

            ai_resp = requests.post(
                'https://openrouter.ai/api/v1/chat/completions',
                headers={'Authorization': f'Bearer {openrouter_key}', 'Content-Type': 'application/json'},
                json={
                    'model': 'openai/gpt-4o-mini',
                    'messages': [
                        {'role': 'system', 'content': system_prompt},
                        {'role': 'user', 'content': user_prompt},
                    ],
                    'max_tokens': 400,
                    'temperature': 0.4,
                },
                timeout=20,
            )
            print('openrouter status=', ai_resp.status_code, 'body=', ai_resp.text[:300])
            if ai_resp.status_code != 200:
                return resp(500, {'error': 'Ошибка AI: ' + ai_resp.text[:200]})

            description = ai_resp.json()['choices'][0]['message']['content'].strip()
            print('success, desc=', description[:100])
            return resp(200, {'description': description, 'web_found': bool(web_context)})

        except Exception as e:
            import traceback
            print('faq-enrich-product error:', traceback.format_exc())
            return resp(500, {'error': str(e) or repr(e)})

    # --- POST ?r=faq-upload  — загрузка картинки в S3, возвращает CDN url
    if r == 'faq-upload' and method == 'POST':
        if not check_auth(hdrs):
            return resp(401, {'error': 'Unauthorized'})
        body = json.loads(body_str)
        data_b64 = body.get('data', '')
        content_type = body.get('content_type', 'image/jpeg')
        ext = content_type.split('/')[-1].replace('jpeg', 'jpg')
        img_bytes = base64.b64decode(data_b64)
        s3 = boto3.client(
            's3',
            endpoint_url='https://bucket.poehali.dev',
            aws_access_key_id=os.environ['AWS_ACCESS_KEY_ID'],
            aws_secret_access_key=os.environ['AWS_SECRET_ACCESS_KEY'],
        )
        file_key = f'faq-images/{uuid.uuid4().hex}.{ext}'
        s3.put_object(Bucket='files', Key=file_key, Body=img_bytes, ContentType=content_type)
        cdn_url = f"https://cdn.poehali.dev/projects/{os.environ['AWS_ACCESS_KEY_ID']}/bucket/{file_key}"
        return resp(200, {'url': cdn_url})

    # ══════════════════════════════════════════════════════════════
    # БЫСТРЫЕ ОТВЕТЫ
    # ══════════════════════════════════════════════════════════════

    # --- GET ?r=questions
    if r == 'questions' and method == 'GET':
        conn = get_conn(); cur = conn.cursor()
        if wl_id:
            _ensure_wl_initialized(conn, cur, wl_id, 'questions')
            cur.execute(f"SELECT id, pattern, answer, active FROM {SCHEMA}.wl_quick_questions WHERE wl_manager_id=%s ORDER BY id", (wl_id,))
        else:
            cur.execute(f"SELECT id, pattern, answer, active FROM {SCHEMA}.ai_quick_questions ORDER BY id")
        rows = cur.fetchall()
        cur.close(); conn.close()
        return resp(200, {'items': [{'id': row[0], 'pattern': row[1], 'answer': row[2], 'active': row[3]} for row in rows]})

    # --- POST ?r=questions
    if r == 'questions' and method == 'POST':
        if not check_auth(hdrs):
            return resp(401, {'error': 'Unauthorized'})
        body = json.loads(body_str)
        pattern = body.get('pattern', '').strip()
        answer = body.get('answer', '').strip()
        if not pattern or not answer:
            return resp(400, {'error': 'pattern and answer required'})
        conn = get_conn(); cur = conn.cursor()
        if wl_id:
            _ensure_wl_initialized(conn, cur, wl_id, 'questions')
            cur.execute(f"INSERT INTO {SCHEMA}.wl_quick_questions (wl_manager_id, pattern, answer, active) VALUES (%s,%s,%s,true) RETURNING id", (wl_id, pattern, answer))
        else:
            cur.execute(f"INSERT INTO {SCHEMA}.ai_quick_questions (pattern, answer, active) VALUES (%s, %s, true) RETURNING id", (pattern, answer))
        new_id = cur.fetchone()[0]
        conn.commit(); cur.close(); conn.close()
        return resp(200, {'id': new_id, 'ok': True})

    # --- PUT ?r=questions&id=X
    if r == 'questions' and method == 'PUT':
        if not check_auth(hdrs):
            return resp(401, {'error': 'Unauthorized'})
        q_id = int(qs.get('id', '0'))
        body = json.loads(body_str)
        conn = get_conn(); cur = conn.cursor()
        if wl_id:
            cur.execute(
                f"UPDATE {SCHEMA}.wl_quick_questions SET pattern=%s, answer=%s, active=%s WHERE id=%s AND wl_manager_id=%s",
                (body.get('pattern',''), body.get('answer',''), body.get('active', True), q_id, wl_id)
            )
        else:
            cur.execute(
                f"UPDATE {SCHEMA}.ai_quick_questions SET pattern=%s, answer=%s, active=%s WHERE id=%s",
                (body.get('pattern',''), body.get('answer',''), body.get('active', True), q_id)
            )
        conn.commit(); cur.close(); conn.close()
        return resp(200, {'ok': True})

    # --- DELETE ?r=questions&id=X
    if r == 'questions' and method == 'DELETE':
        if not check_auth(hdrs):
            return resp(401, {'error': 'Unauthorized'})
        q_id = int(qs.get('id', '0'))
        conn = get_conn(); cur = conn.cursor()
        if wl_id:
            cur.execute(f"DELETE FROM {SCHEMA}.wl_quick_questions WHERE id=%s AND wl_manager_id=%s", (q_id, wl_id))
        else:
            cur.execute(f"DELETE FROM {SCHEMA}.ai_quick_questions WHERE id = %s", (q_id,))
        conn.commit(); cur.close(); conn.close()
        return resp(200, {'ok': True})

    # ══════════════════════════════════════════════════════════════
    # НАСТРОЙКИ КАТЕГОРИЙ
    # ══════════════════════════════════════════════════════════════

    # --- GET ?r=category_settings
    if r == 'category_settings' and method == 'GET':
        conn = get_conn(); cur = conn.cursor()
        # Глобальные
        cur.execute(f"SELECT category, is_material, category_rule, is_wall_item, show_in_drum, use_installation_price FROM {SCHEMA}.price_category_settings ORDER BY category")
        global_rows = {row[0]: {'category': row[0], 'is_material': row[1], 'category_rule': row[2] or '', 'is_wall_item': row[3] if row[3] is not None else True, 'show_in_drum': row[4] if row[4] is not None else True, 'use_installation_price': bool(row[5]) if row[5] is not None else False} for row in cur.fetchall()}
        if wl_id:
            # Накладываем WL-переопределения
            cur.execute(f"SELECT category, is_material, category_rule, is_wall_item, show_in_drum FROM {SCHEMA}.wl_category_settings WHERE wl_manager_id=%s", (wl_id,))
            for row in cur.fetchall():
                global_rows[row[0]] = {'category': row[0], 'is_material': row[1], 'category_rule': row[2] or '', 'is_wall_item': row[3] if row[3] is not None else True, 'show_in_drum': row[4] if row[4] is not None else True, 'use_installation_price': global_rows.get(row[0], {}).get('use_installation_price', False)}
        cur.close(); conn.close()
        return resp(200, {'items': list(global_rows.values())})

    # --- PUT ?r=category_settings
    if r == 'category_settings' and method == 'PUT':
        body = json.loads(body_str)
        category = body.get('category', '').strip()
        if not category:
            return resp(400, {'error': 'category required'})
        # Читаем только переданные поля; None = не передано, не перезаписываем
        is_material  = bool(body['is_material'])  if 'is_material'  in body else None
        is_wall_item = bool(body['is_wall_item']) if 'is_wall_item' in body else None
        show_in_drum = bool(body['show_in_drum']) if 'show_in_drum' in body else None
        category_rule = body['category_rule'].strip() if 'category_rule' in body else None
        use_installation_price = bool(body['use_installation_price']) if 'use_installation_price' in body else None

        conn = get_conn(); cur = conn.cursor()
        table = f"{SCHEMA}.wl_category_settings" if wl_id else f"{SCHEMA}.price_category_settings"

        # Строим динамический UPDATE только для переданных полей
        update_parts = []
        params = []
        if is_material  is not None: update_parts.append("is_material=%s");  params.append(is_material)
        if is_wall_item is not None: update_parts.append("is_wall_item=%s"); params.append(is_wall_item)
        if show_in_drum is not None: update_parts.append("show_in_drum=%s"); params.append(show_in_drum)
        if category_rule is not None: update_parts.append("category_rule=%s"); params.append(category_rule)
        if use_installation_price is not None and not wl_id:
            update_parts.append("use_installation_price=%s"); params.append(use_installation_price)
        update_parts.append("updated_at=now()")

        if wl_id:
            # INSERT с дефолтами, ON CONFLICT обновляет только переданные поля
            ins_cols = "wl_manager_id, category"
            ins_vals = [wl_id, category]
            if is_material  is not None: ins_cols += ", is_material";  ins_vals.append(is_material)
            if is_wall_item is not None: ins_cols += ", is_wall_item"; ins_vals.append(is_wall_item)
            if show_in_drum is not None: ins_cols += ", show_in_drum"; ins_vals.append(show_in_drum)
            if category_rule is not None: ins_cols += ", category_rule"; ins_vals.append(category_rule)
            placeholders = ", ".join(["%s"] * len(ins_vals))
            upd = ", ".join(update_parts)
            cur.execute(
                f"INSERT INTO {table} ({ins_cols}) VALUES ({placeholders}) ON CONFLICT (wl_manager_id, category) DO UPDATE SET {upd}",
                ins_vals + params
            )
        else:
            ins_cols = "category"
            ins_vals = [category]
            if is_material  is not None: ins_cols += ", is_material";  ins_vals.append(is_material)
            if is_wall_item is not None: ins_cols += ", is_wall_item"; ins_vals.append(is_wall_item)
            if show_in_drum is not None: ins_cols += ", show_in_drum"; ins_vals.append(show_in_drum)
            if category_rule is not None: ins_cols += ", category_rule"; ins_vals.append(category_rule)
            if use_installation_price is not None: ins_cols += ", use_installation_price"; ins_vals.append(use_installation_price)
            placeholders = ", ".join(["%s"] * len(ins_vals))
            upd = ", ".join(update_parts)
            cur.execute(
                f"INSERT INTO {table} ({ins_cols}) VALUES ({placeholders}) ON CONFLICT (category) DO UPDATE SET {upd}",
                ins_vals + params
            )
        conn.commit(); cur.close(); conn.close()
        return resp(200, {'ok': True})

    # ══════════════════════════════════════════════════════════════
    # ПРАЙС
    # ══════════════════════════════════════════════════════════════

    # --- GET ?r=prices
    if r == 'prices' and method == 'GET':
        conn = get_conn(); cur = conn.cursor()
        cur.execute(f"SELECT id, category, name, price, unit, description, sort_order, active, calc_rule, bundle, synonyms, when_condition, when_not_condition, client_changes, purchase_price, image_url, category_image_url, installation_price, measure_price, management_price FROM {SCHEMA}.ai_prices ORDER BY sort_order, id")
        rows = cur.fetchall()
        cur.execute(f"SELECT category, is_material, is_wall_item, show_in_drum, use_installation_price FROM {SCHEMA}.price_category_settings")
        cat_settings = {row[0]: {'is_material': row[1], 'is_wall_item': row[2] if row[2] is not None else True, 'show_in_drum': row[3] if row[3] is not None else True, 'use_installation_price': bool(row[4]) if row[4] is not None else False} for row in cur.fetchall()}

        wl_overrides = {}
        if wl_id:
            cur.execute(f"SELECT price_id, price, purchase_price, active, description, synonyms, image_url, category_image_url FROM {SCHEMA}.wl_price_overrides WHERE wl_manager_id=%s", (wl_id,))
            for ov in cur.fetchall():
                wl_overrides[ov[0]] = {
                    'price': ov[1], 'purchase_price': ov[2], 'active': ov[3],
                    'description': ov[4], 'synonyms': ov[5],
                    'image_url': ov[6], 'category_image_url': ov[7],
                }
            # WL category_settings тоже
            cur.execute(f"SELECT category, is_material, is_wall_item, show_in_drum FROM {SCHEMA}.wl_category_settings WHERE wl_manager_id=%s", (wl_id,))
            for row in cur.fetchall():
                cat_settings[row[0]] = {'is_material': row[1], 'is_wall_item': row[2] if row[2] is not None else True, 'show_in_drum': row[3] if row[3] is not None else True}

        cur.close(); conn.close()

        items = []
        for row in rows:
            pid = row[0]
            ov = wl_overrides.get(pid, {})
            cs = cat_settings.get(row[1], {})
            items.append({
                'id': pid,
                'category': row[1],
                'name': row[2],
                'price': ov['price'] if ov.get('price') is not None else row[3],
                'unit': row[4],
                'description': ov['description'] if ov.get('description') is not None else (row[5] or ''),
                'sort_order': row[6],
                'active': ov['active'] if ov.get('active') is not None else row[7],
                'calc_rule': row[8] or '',
                'bundle': row[9] or '[]',
                'synonyms': ov['synonyms'] if ov.get('synonyms') is not None else (row[10] or ''),
                'when_condition': row[11] or '',
                'when_not_condition': row[12] or '',
                'client_changes': row[13] or '',
                'purchase_price': ov['purchase_price'] if ov.get('purchase_price') is not None else (row[14] or 0),
                'installation_price': row[17] if row[17] is not None else 100,
                'measure_price': row[18] if row[18] is not None else 100,
                'management_price': row[19] if row[19] is not None else 100,
                'is_material': cs.get('is_material', True),
                'is_wall_item': cs.get('is_wall_item', True),
                'show_in_drum': cs.get('show_in_drum', True),
                'use_installation_price': cs.get('use_installation_price', False),
                'image_url': ov['image_url'] if ov.get('image_url') is not None else row[15],
                'category_image_url': ov['category_image_url'] if ov.get('category_image_url') is not None else row[16],
            })
        return resp(200, {'items': items})

    # --- POST ?r=prices  (добавление позиции — только мастер)
    if r == 'prices' and method == 'POST':
        if not check_auth(hdrs):
            return resp(401, {'error': 'Unauthorized'})
        if wl_id:
            return resp(403, {'error': 'WL-клиент не может добавлять позиции прайса'})
        body = json.loads(body_str)
        category = body.get('category', '').strip()
        name = body.get('name', '').strip()
        if not category or not name:
            return resp(400, {'error': 'category and name required'})
        conn = get_conn(); cur = conn.cursor()
        cur.execute(f"SELECT COALESCE(MAX(sort_order), 0) + 10 FROM {SCHEMA}.ai_prices WHERE category = %s", (category,))
        sort_order = cur.fetchone()[0]
        synonyms = body.get('synonyms', '')
        cur.execute(
            f"INSERT INTO {SCHEMA}.ai_prices (category, name, price, purchase_price, unit, description, sort_order, active, synonyms) VALUES (%s,%s,%s,%s,%s,%s,%s,true,%s) RETURNING id",
            (category, name, int(body.get('price', 0)), int(body.get('purchase_price', 0)), body.get('unit', 'шт'), body.get('description', ''), sort_order, synonyms)
        )
        new_id = cur.fetchone()[0]
        _save_complex_exceptions(conn, cur, synonyms, name)
        cur2 = conn.cursor()
        cur2.execute(
            f"INSERT INTO {SCHEMA}.price_category_settings (category, is_material, category_rule, updated_at) VALUES (%s, true, '', now()) ON CONFLICT (category) DO NOTHING",
            (category,)
        )
        conn.commit()
        cur2.close(); cur.close(); conn.close()
        return resp(200, {'id': new_id, 'ok': True})

    # --- PUT ?r=prices&id=X  (обновление позиции)
    if r == 'prices' and method == 'PUT' and 'rename_category' not in qs:
        if not check_auth(hdrs):
            return resp(401, {'error': 'Unauthorized'})
        price_id = int(qs.get('id', '0'))
        body = json.loads(body_str)
        conn = get_conn(); cur = conn.cursor()

        if wl_id:
            # WL: сохраняем только изменённые поля в override-таблицу
            ov_price = body.get('price')
            ov_purchase = body.get('purchase_price')
            ov_active = body.get('active')
            ov_desc = body.get('description')
            ov_syn = body.get('synonyms')
            cur.execute(
                f"""INSERT INTO {SCHEMA}.wl_price_overrides (wl_manager_id, price_id, price, purchase_price, active, description, synonyms, updated_at)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,now())
                    ON CONFLICT (wl_manager_id, price_id) DO UPDATE SET
                        price=COALESCE(EXCLUDED.price, {SCHEMA}.wl_price_overrides.price),
                        purchase_price=COALESCE(EXCLUDED.purchase_price, {SCHEMA}.wl_price_overrides.purchase_price),
                        active=COALESCE(EXCLUDED.active, {SCHEMA}.wl_price_overrides.active),
                        description=COALESCE(EXCLUDED.description, {SCHEMA}.wl_price_overrides.description),
                        synonyms=COALESCE(EXCLUDED.synonyms, {SCHEMA}.wl_price_overrides.synonyms),
                        updated_at=now()""",
                (wl_id, price_id,
                 int(ov_price) if ov_price is not None else None,
                 int(ov_purchase) if ov_purchase is not None else None,
                 ov_active, ov_desc, ov_syn)
            )
        else:
            # Мастер: обновляет глобальный прайс
            synonyms = body.get('synonyms', '')
            sort_order = body.get('sort_order')
            when_condition = body.get('when_condition', '')
            when_not_condition = body.get('when_not_condition', '')
            client_changes = body.get('client_changes', '')
            purchase_price = int(body.get('purchase_price', 0))
            installation_price = int(body.get('installation_price', 100))
            measure_price = int(body.get('measure_price', 100))
            management_price = int(body.get('management_price', 100))
            if sort_order is not None:
                cur.execute(
                    f"UPDATE {SCHEMA}.ai_prices SET name=%s, price=%s, purchase_price=%s, installation_price=%s, measure_price=%s, management_price=%s, unit=%s, description=%s, active=%s, calc_rule=%s, bundle=%s, synonyms=%s, when_condition=%s, when_not_condition=%s, client_changes=%s, sort_order=%s, updated_at=now() WHERE id=%s",
                    (body.get('name',''), int(body.get('price', 0)), purchase_price, installation_price, measure_price, management_price, body.get('unit',''), body.get('description',''), body.get('active', True), body.get('calc_rule',''), body.get('bundle','[]'), synonyms, when_condition, when_not_condition, client_changes, int(sort_order), price_id)
                )
            else:
                cur.execute(
                    f"UPDATE {SCHEMA}.ai_prices SET name=%s, price=%s, purchase_price=%s, installation_price=%s, measure_price=%s, management_price=%s, unit=%s, description=%s, active=%s, calc_rule=%s, bundle=%s, synonyms=%s, when_condition=%s, when_not_condition=%s, client_changes=%s, updated_at=now() WHERE id=%s",
                    (body.get('name',''), int(body.get('price', 0)), purchase_price, installation_price, measure_price, management_price, body.get('unit',''), body.get('description',''), body.get('active', True), body.get('calc_rule',''), body.get('bundle','[]'), synonyms, when_condition, when_not_condition, client_changes, price_id)
                )
            _save_complex_exceptions(conn, cur, synonyms, body.get('name', ''))

        conn.commit(); cur.close(); conn.close()
        return resp(200, {'ok': True})

    # --- PUT ?r=prices&rename_category  (переименование категории — только мастер)
    if r == 'prices' and method == 'PUT' and 'rename_category' in qs:
        if not check_auth(hdrs):
            return resp(401, {'error': 'Unauthorized'})
        if wl_id:
            return resp(403, {'error': 'WL-клиент не может переименовывать категории'})
        body = json.loads(body_str)
        old_name = body.get('old_name', '').strip()
        new_name = body.get('new_name', '').strip()
        if not old_name or not new_name:
            return resp(400, {'error': 'old_name and new_name required'})
        conn = get_conn(); cur = conn.cursor()
        cur.execute(f"UPDATE {SCHEMA}.ai_prices SET category=%s WHERE category=%s", (new_name, old_name))
        conn.commit(); cur.close(); conn.close()
        return resp(200, {'ok': True})

    # --- DELETE ?r=prices&id=X  (удаление — только мастер)
    if r == 'prices' and method == 'DELETE':
        if not check_auth(hdrs):
            return resp(401, {'error': 'Unauthorized'})
        if wl_id:
            return resp(403, {'error': 'WL-клиент не может удалять позиции прайса'})
        price_id = int(qs.get('id', '0'))
        conn = get_conn(); cur = conn.cursor()
        cur.execute(f"DELETE FROM {SCHEMA}.ai_prices WHERE id = %s", (price_id,))
        conn.commit(); cur.close(); conn.close()
        return resp(200, {'ok': True})

    # ══════════════════════════════════════════════════════════════
    # КОРРЕКЦИИ (уже изолированы по company_id, не меняем)
    # ══════════════════════════════════════════════════════════════

    if r == 'corrections' and method == 'GET':
        user_id, user_role = get_company_id_from_token(hdrs)
        is_master_corr = check_auth(hdrs) and not wl_id
        conn = get_conn(); cur = conn.cursor()
        def fmt_dt(v):
            if v is None: return None
            s = str(v)
            return s if '+' in s or s.endswith('Z') else s + '+00:00'
        if is_master_corr:
            cur.execute(f"SELECT id, session_id, user_text, recognized_json, corrected_json, status, created_at, suggested_items, llm_answer FROM {SCHEMA}.bot_corrections ORDER BY created_at DESC LIMIT 500")
        elif user_id and user_role in ('company', 'installer'):
            cur.execute(f"SELECT id, session_id, user_text, recognized_json, corrected_json, status, created_at, suggested_items, llm_answer FROM {SCHEMA}.bot_corrections WHERE company_id=%s ORDER BY created_at DESC LIMIT 200", (user_id,))
        elif wl_id:
            cur.execute(f"SELECT id, session_id, user_text, recognized_json, corrected_json, status, created_at, suggested_items, llm_answer FROM {SCHEMA}.bot_corrections WHERE company_id=%s ORDER BY created_at DESC LIMIT 200", (wl_id,))
        else:
            cur.close(); conn.close()
            return resp(401, {'error': 'Unauthorized'})
        rows = cur.fetchall()
        cur.close(); conn.close()
        return resp(200, {'items': [{'id': row[0], 'session_id': row[1] or '', 'user_text': row[2], 'recognized_json': row[3], 'corrected_json': row[4], 'status': row[5], 'created_at': fmt_dt(row[6]), 'suggested_items': row[7], 'llm_answer': row[8]} for row in rows]})

    if r == 'corrections' and method == 'PUT':
        if not check_auth(hdrs):
            return resp(401, {'error': 'Unauthorized'})
        body = json.loads(body_str)
        corr_id = int(qs.get('id', body.get('id', 0)))
        status = body.get('status', 'pending')
        corrected = body.get('corrected_json')
        conn = get_conn(); cur = conn.cursor()
        cur.execute(
            f"UPDATE {SCHEMA}.bot_corrections SET status=%s, corrected_json=%s, reviewed_at=now() WHERE id=%s",
            (status, json.dumps(corrected, ensure_ascii=False) if corrected else None, corr_id)
        )
        conn.commit(); cur.close(); conn.close()
        return resp(200, {'ok': True})

    if r == 'corrections' and method == 'DELETE':
        if not check_auth(hdrs):
            return resp(401, {'error': 'Unauthorized'})
        body = json.loads(body_str)
        corr_id = int(qs.get('id', body.get('id', 0)))
        conn = get_conn(); cur = conn.cursor()
        cur.execute(f"DELETE FROM {SCHEMA}.bot_corrections WHERE id = %s", (corr_id,))
        conn.commit(); cur.close(); conn.close()
        return resp(200, {'ok': True})

    # ══════════════════════════════════════════════════════════════
    # СТОП-СЛОВА (только мастер)
    # ══════════════════════════════════════════════════════════════

    if r == 'stop-words' and method == 'GET':
        conn = get_conn(); cur = conn.cursor()
        cur.execute(f"SELECT id, word, created_at FROM {SCHEMA}.stop_words ORDER BY created_at DESC")
        rows = cur.fetchall()
        cur.close(); conn.close()
        return resp(200, {'items': [{'id': r[0], 'word': r[1], 'created_at': str(r[2])} for r in rows]})

    if r == 'stop-words' and method == 'POST':
        if not check_auth(hdrs):
            return resp(401, {'error': 'Unauthorized'})
        body = json.loads(body_str)
        word = body.get('word', '').strip().lower()
        if not word:
            return resp(400, {'error': 'word required'})
        conn = get_conn(); cur = conn.cursor()
        cur.execute(f"INSERT INTO {SCHEMA}.stop_words (word) VALUES (%s) ON CONFLICT DO NOTHING", (word,))
        conn.commit(); cur.close(); conn.close()
        return resp(200, {'ok': True})

    if r == 'stop-words' and method == 'DELETE':
        if not check_auth(hdrs):
            return resp(401, {'error': 'Unauthorized'})
        word_id = int(qs.get('id', '0'))
        conn = get_conn(); cur = conn.cursor()
        cur.execute(f"DELETE FROM {SCHEMA}.stop_words WHERE id = %s", (word_id,))
        conn.commit(); cur.close(); conn.close()
        return resp(200, {'ok': True})

    # ══════════════════════════════════════════════════════════════
    # ТИПЫ И ЗНАЧЕНИЯ ПРАВИЛ (только мастер, WL только читает)
    # ══════════════════════════════════════════════════════════════

    if r == 'rule-types' and method == 'GET':
        conn = get_conn(); cur = conn.cursor()
        cur.execute(f"SELECT id, name, label, description, placeholder, sort_order, active FROM {SCHEMA}.ai_rule_types ORDER BY sort_order, id")
        rows = cur.fetchall()
        cur.close(); conn.close()
        return resp(200, {'items': [{'id': r[0], 'name': r[1], 'label': r[2], 'description': r[3], 'placeholder': r[4], 'sort_order': r[5], 'active': r[6]} for r in rows]})

    if r == 'rule-types' and method == 'POST':
        if not check_auth(hdrs) or wl_id:
            return resp(403, {'error': 'Только мастер может создавать типы правил'})
        body = json.loads(body_str)
        label = body.get('label', '').strip()
        name = body.get('name', '').strip() or label.lower().replace(' ', '_').replace('-', '_')[:30]
        if not label:
            return resp(400, {'error': 'label required'})
        conn = get_conn(); cur = conn.cursor()
        cur.execute(f"SELECT COALESCE(MAX(sort_order), 0) + 10 FROM {SCHEMA}.ai_rule_types")
        sort_order = cur.fetchone()[0]
        cur.execute(
            f"INSERT INTO {SCHEMA}.ai_rule_types (name, label, description, placeholder, sort_order) VALUES (%s,%s,%s,%s,%s) RETURNING id",
            (name, label, body.get('description', ''), body.get('placeholder', ''), sort_order)
        )
        new_id = cur.fetchone()[0]
        conn.commit(); cur.close(); conn.close()
        return resp(200, {'id': new_id, 'ok': True})

    if r == 'rule-types' and method == 'PUT':
        if not check_auth(hdrs) or wl_id:
            return resp(403, {'error': 'Только мастер может изменять типы правил'})
        rule_id = int(qs.get('id', '0'))
        body = json.loads(body_str)
        conn = get_conn(); cur = conn.cursor()
        cur.execute(
            f"UPDATE {SCHEMA}.ai_rule_types SET label=%s, description=%s, placeholder=%s, active=%s WHERE id=%s",
            (body.get('label', ''), body.get('description', ''), body.get('placeholder', ''), body.get('active', True), rule_id)
        )
        conn.commit(); cur.close(); conn.close()
        return resp(200, {'ok': True})

    if r == 'rule-types' and method == 'DELETE':
        if not check_auth(hdrs) or wl_id:
            return resp(403, {'error': 'Только мастер может удалять типы правил'})
        rule_id = int(qs.get('id', '0'))
        conn = get_conn(); cur = conn.cursor()
        cur.execute(f"DELETE FROM {SCHEMA}.ai_rule_values WHERE rule_type_id = %s", (rule_id,))
        cur.execute(f"DELETE FROM {SCHEMA}.ai_rule_types WHERE id = %s", (rule_id,))
        conn.commit(); cur.close(); conn.close()
        return resp(200, {'ok': True})

    if r == 'rule-values' and method == 'GET':
        price_id = int(qs.get('price_id', '0'))
        conn = get_conn(); cur = conn.cursor()
        cur.execute(f"SELECT rule_type_id, value FROM {SCHEMA}.ai_rule_values WHERE price_id = %s", (price_id,))
        rows = cur.fetchall()
        cur.close(); conn.close()
        return resp(200, {'values': {str(r[0]): r[1] for r in rows}})

    if r == 'rule-values' and method == 'POST':
        if not check_auth(hdrs) or wl_id:
            return resp(403, {'error': 'Только мастер может изменять значения правил'})
        body = json.loads(body_str)
        price_id = int(body.get('price_id', 0))
        rule_type_id = int(body.get('rule_type_id', 0))
        value = body.get('value', '')
        conn = get_conn(); cur = conn.cursor()
        cur.execute(
            f"INSERT INTO {SCHEMA}.ai_rule_values (price_id, rule_type_id, value, updated_at) VALUES (%s,%s,%s,now()) ON CONFLICT (price_id, rule_type_id) DO UPDATE SET value=%s, updated_at=now()",
            (price_id, rule_type_id, value, value)
        )
        conn.commit(); cur.close(); conn.close()
        return resp(200, {'ok': True})

    # ══════════════════════════════════════════════════════════════
    # НАСТРОЙКИ AI (только мастер)
    # ══════════════════════════════════════════════════════════════

    if r == 'settings' and method == 'GET':
        conn = get_conn(); cur = conn.cursor()
        cur.execute(f"SELECT key, value FROM {SCHEMA}.ai_settings")
        rows_s = cur.fetchall()
        cur.close(); conn.close()
        return resp(200, {row[0]: row[1] for row in rows_s})

    if r == 'settings' and method == 'PUT':
        if not check_auth(hdrs):
            return resp(401, {'error': 'Unauthorized'})
        body = json.loads(body_str)
        conn = get_conn(); cur = conn.cursor()
        for key, value in body.items():
            cur.execute(
                f"INSERT INTO {SCHEMA}.ai_settings (key, value, updated_at) VALUES (%s, %s, now()) ON CONFLICT (key) DO UPDATE SET value=%s, updated_at=now()",
                (key, str(value), str(value))
            )
        conn.commit(); cur.close(); conn.close()
        return resp(200, {'ok': True})

    # ══════════════════════════════════════════════════════════════
    # ПОДБОР СИНОНИМОВ / ГЕНЕРАЦИЯ (мастер и WL могут использовать)
    # ══════════════════════════════════════════════════════════════

    if r == 'match-synonym' and method == 'POST':
        if not check_auth(hdrs):
            return resp(401, {'error': 'Unauthorized'})
        body = json.loads(body_str)
        word = body.get('word', '').strip()
        prices_list = body.get('prices', [])
        if not word:
            return resp(400, {'error': 'word required'})

        openrouter_key = os.environ.get('OPENROUTER_API_KEY_2', '')
        if not openrouter_key:
            return resp(500, {'error': 'No LLM key'})

        if word.startswith('GENERATE_DESCRIPTION:'):
            parts = word[len('GENERATE_DESCRIPTION:'):].split('|')
            item_name = parts[0].strip() if parts else word
            item_cat  = parts[1].strip() if len(parts) > 1 else ''
            desc_prompt = f"""Ты эксперт по натяжным потолкам.
Напиши короткое описание (1 предложение, максимум 10 слов) для позиции прайса: «{item_name}» (категория: {item_cat}).
Описание должно объяснять что это такое — для понимания AI-бота, не для клиента.
Отвечай ТОЛЬКО описанием, без кавычек и пояснений:"""
            try:
                llm_r = requests.post(
                    'https://openrouter.ai/api/v1/chat/completions',
                    json={'model': 'openai/gpt-4o-mini', 'messages': [{'role': 'user', 'content': desc_prompt}], 'max_tokens': 60, 'temperature': 0.5},
                    headers={'Authorization': f'Bearer {openrouter_key}', 'Content-Type': 'application/json'},
                    timeout=20,
                )
                if llm_r.status_code == 200:
                    description = llm_r.json()['choices'][0]['message']['content'].strip().strip('"').strip("'")
                    return resp(200, {'description': description})
            except Exception as e:
                return resp(500, {'error': str(e)})
            return resp(500, {'error': 'LLM failed'})

        if word.startswith('GENERATE_SYNONYMS:'):
            parts = word[len('GENERATE_SYNONYMS:'):].split('|')
            item_name = parts[0].strip() if parts else word
            item_desc = parts[1].strip() if len(parts) > 1 else ''
            item_cat  = parts[2].strip() if len(parts) > 2 else ''
            syn_prompt = f"""Ты эксперт по натяжным потолкам. 
Позиция прайса: «{item_name}» (категория: {item_cat})
{f'Описание: {item_desc}' if item_desc else ''}

Придумай 5-10 синонимов — разные способы как клиент может написать эту позицию в запросе.
Например сокращения, разговорные формы, ошибочные написания, варианты с числами.

Отвечай ТОЛЬКО через запятую, без пояснений:"""
            try:
                llm_r = requests.post(
                    'https://openrouter.ai/api/v1/chat/completions',
                    json={'model': 'openai/gpt-4o-mini', 'messages': [{'role': 'user', 'content': syn_prompt}], 'max_tokens': 150, 'temperature': 0.7},
                    headers={'Authorization': f'Bearer {openrouter_key}', 'Content-Type': 'application/json'},
                    timeout=20,
                )
                if llm_r.status_code == 200:
                    synonyms = llm_r.json()['choices'][0]['message']['content'].strip()
                    return resp(200, {'synonyms': synonyms})
            except Exception as e:
                return resp(500, {'error': str(e)})
            return resp(500, {'error': 'LLM failed'})

        if not prices_list:
            return resp(400, {'error': 'prices required'})

        prices_text = '\n'.join(
            f"{p['id']}. {p['name']} ({p['category']})"
            + (f" [синонимы: {p['synonyms']}]" if p.get('synonyms') else '')
            for p in prices_list[:80]
        )
        prompt = f"""Ты помощник сметчика натяжных потолков. 
Тебе дано слово/фраза из запроса клиента: «{word}»

Из списка позиций прайса найди ОДНУ наиболее подходящую позицию для этого слова.
Отвечай ТОЛЬКО числом — ID позиции. Если ни одна не подходит — ответь 0.

Список позиций:
{prices_text}

Ответ (только число):"""
        try:
            llm_resp = requests.post(
                'https://openrouter.ai/api/v1/chat/completions',
                json={'model': 'openai/gpt-4o-mini', 'messages': [{'role': 'user', 'content': prompt}], 'max_tokens': 10, 'temperature': 0},
                headers={'Authorization': f'Bearer {openrouter_key}', 'Content-Type': 'application/json'},
                timeout=15,
            )
            if llm_resp.status_code != 200:
                return resp(500, {'error': f'LLM error {llm_resp.status_code}'})
            content = llm_resp.json()['choices'][0]['message']['content'].strip()
            matched_id = int(''.join(filter(str.isdigit, content)) or '0')
            return resp(200, {'matched_id': matched_id})
        except Exception as e:
            return resp(500, {'error': str(e)})

    # ══════════════════════════════════════════════════════════════
    # XLSX IMPORT (только мастер)
    # ══════════════════════════════════════════════════════════════

    action = qs.get('action', '')
    if action in ('read', 'import'):
        items, headers_list = load_xlsx()
        if not items:
            return resp(200, {'total': 0, 'items': []})
        if action == 'read':
            page = int(qs.get('page', '0'))
            per_page = int(qs.get('per_page', '50'))
            start = page * per_page
            return resp(200, {'total': len(items), 'page': page, 'headers': headers_list, 'items': items[start:start+per_page]})
        if action == 'import':
            if not check_auth(hdrs):
                return resp(401, {'error': 'Unauthorized'})
            conn = get_conn(); cur = conn.cursor()
            cur.execute(f"TRUNCATE TABLE {SCHEMA}.faq_items RESTART IDENTITY")
            imported = 0
            for item in items:
                title   = item.get('title', item.get('заголовок', item.get('название', '')))
                s_title = item.get('search title', item.get('search_title', item.get('поиск', title)))
                content = item.get('content', item.get('содержимое', item.get('описание', '')))
                used    = str(item.get('used', item.get('активен', 'True'))).strip().lower() not in ('false', '0', 'нет', 'no', '')
                if not title and not content:
                    continue
                cur.execute(f"INSERT INTO {SCHEMA}.faq_items (title, search_title, content, used) VALUES (%s, %s, %s, %s)",
                            (title, s_title or title, content, used))
                imported += 1
            conn.commit(); cur.close(); conn.close()
            return resp(200, {'ok': True, 'imported': imported})

    return resp(400, {'error': 'unknown resource. Use ?r=prompt|faq|questions|prices|login'})